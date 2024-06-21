import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Load the spreadsheet file
workbook = load_workbook(filename="orders.xlsx")
sheet = workbook.active

# Get the last order ID
last_order_id = 0
if sheet.max_row > 1:
    cell_value = sheet.cell(row=sheet.max_row, column=1).value
    if cell_value is not None:
        last_order_id = int(cell_value)

# Create a Streamlit app
st.title("Order App")

# Create a form to take orders
with st.form("order_form"):
    customer_name = st.text_input("Customer Name")
    customer_number = st.text_input("Customer Number")
    menu_items = [
        "Nippatt Masala (Rs. 35)",
        "Peanut Masala (Rs. 40)",
        "Girmit (Rs. 25)",
        "Cornflakes Masala (Rs. 40)",
        "Bun Mirchi (Rs. 25)",
        "Paneer Corn Tikka (Rs. 70)",
        "Italian Cheese Corn (Rs. 70)",
        "Mexican Cheese Corn (Rs. 70)",
        "South Style Corn (Rs. 80)",
        "Spice Lemon Chilli Corn (Rs. 80)",
        "Chipotle Style Corn (Rs. 70)",
        "Barbie Corn (Rs. 70)",
        "Corn Salsa (Rs. 70)",
        "Spicy Corn Siraja (Rs. 70)",
        "Corn Island (Rs. 70)",
        "Jalapeno Corn (Rs. 70)",
        "Shezwan Corn (Rs. 70)",
        "Jamaican Jerk Corn (Rs. 70)",
        "Corn Nachos (Rs. 70)"
    ]
    prices = {
        "Nippatt Masala": 35,
        "Peanut Masala": 40,
        "Girmit": 25,
        "Cornflakes Masala": 40,
        "Bun Mirchi": 25,
        "Paneer Corn Tikka": 70,
        "Italian Cheese Corn": 70,
        "Mexican Cheese Corn": 70,
        "South Style Corn": 80,
        "Spice Lemon Chilli Corn": 80,
        "Chipotle Style Corn": 70,
        "Barbie Corn": 70,
        "Corn Salsa": 70,
        "Spicy Corn Siraja": 70,
        "Corn Island": 70,
        "Jalapeno Corn": 70,
        "Shezwan Corn": 70,
        "Jamaican Jerk Corn": 70,
        "Corn Nachos": 70
    }
    selected_menu = st.multiselect("Select Menu Items", menu_items)
    paid_by_cash = st.checkbox("Paid by Cash")
    paid_by_upi = st.checkbox("Paid by UPI")
    submitted = st.form_submit_button("Submit")

    if submitted:
        # Auto-increment the order ID
        order_id = last_order_id + 1

        # Calculate the total amount and prepare the order row
        total_amount = 0
        order_row = [order_id, customer_name, customer_number]

        for item in menu_items:
            item_name = item.split(" (")[0]
            if item in selected_menu:
                order_row.append(item_name)
                total_amount += prices[item_name]
            else:
                order_row.append("")

        # Append the total amount and payment method to the order row
        order_row.append(total_amount)
        payment_method = "Cash" if paid_by_cash else "UPI" if paid_by_upi else "None"
        order_row.append(payment_method)

        # Add the order to the spreadsheet
        sheet.append(order_row)
        workbook.save("orders.xlsx")
        st.success(f"Order submitted successfully! Total amount: Rs. {total_amount:.2f}")

# Create a page to show orders
st.header("Orders")
orders = []
for row in sheet.iter_rows(min_row=1, values_only=True):
    orders.append(list(row))
columns = ["Order ID", "Customer Name", "Customer Number"] + [item.split(" (")[0] for item in menu_items] + ["Total Amount", "Payment Method"]
orders_df = pd.DataFrame(orders, columns=columns)
st.write(orders_df)
